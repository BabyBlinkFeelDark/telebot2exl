from datetime import datetime
import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def fetch_courier_data_to_excel(st_p, end_p, host, port, user, password, database, output_file):

    today = datetime.now()
    # Форматируем дату в нужный формат
    formatted_date = today.strftime('%Y-%m-%d')
    """
    Подключается к базе данных, извлекает данные из таблицы courier и экспортирует их в файл Excel.

    :param st_p: Начало периода.
    :param end_p: Конец периода.
    :param host: Хост базы данных.
    :param port: Порт базы данных.
    :param user: Имя пользователя для подключения.
    :param password: Пароль для подключения.
    :param database: Имя базы данных.
    :param output_file: Путь для сохранения Excel-файла.
    """
    try:
        # Подключение к базе данных
        connection = psycopg2.connect(
            host=host,
            port=port,
            user=user,
            password=password,
            database=database
        )
        print("Успешное подключение к базе данных.")

        # SQL-запрос
        query = (f"""SELECT
                    couriers.courier_id AS ID,
                    couriers.courier_name AS ФИО,
                    COUNT(orders.order_id) AS КОЛИЧЕСТВО_ДОСТАВОК,
                    SUM(orders.time_taken) AS ОБЩЕЕ_ВРЕМЯ_ОЖИДАНИЯ,
                    ROUND(SUM(orders.time_taken)::numeric / NULLIF(COUNT(orders.order_id), 0), 2) AS СРЕДНЕЕ_ВРЕМЯ_ОЖИДАНИЯ
                FROM couriers
                JOIN orders ON orders.courier_id = couriers.courier_id
                WHERE orders.cur_time >= '{formatted_date} {st_p:02d}:00:00' AND orders.cur_time < '{formatted_date} {end_p:02d}:00:00'
                GROUP BY couriers.courier_id, couriers.courier_name
                ORDER BY СРЕДНЕЕ_ВРЕМЯ_ОЖИДАНИЯ DESC;""")

    # Используем pandas.read_sql_query с pgdb
        cursor = connection.cursor()
        cursor.execute(query)

        # Извлекаем данные и преобразуем в DataFrame
        columns = [desc[0] for desc in cursor.description]
        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=columns)

        # Сохранение в Excel
        df.to_excel(output_file, index=False)
        print(f"Данные успешно экспортированы в файл {output_file}.")

        # Открываем файл для редактирования
        wb = load_workbook(output_file)
        ws = wb.active

        # Устанавливаем ширину определенных колонок (например, 1 и 3)
        columns_to_adjust = [1, 2, 3, 4, 5]  # Измените на нужные вам номера столбцов
        new_widths = [5, 20, 10, 10, 10]  # Новые ширины для указанных столбцов

        for col, width in zip(columns_to_adjust, new_widths):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width  # Устанавливаем ширину в 20

        # Сохраняем изменения
        wb.save(output_file)
        print(f"Данные успешно экспортированы в файл {output_file}.")

    except Exception as e:
        print(f"Ошибка: {e}")

    finally:
        # Закрытие подключения
        if 'connection' in locals() and connection:
            connection.close()
            print("Подключение к базе данных закрыто.")
